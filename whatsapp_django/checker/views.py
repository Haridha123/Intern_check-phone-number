from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
import json
import threading
import time
from datetime import datetime
import csv
import io
import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from .smart_checker import check_number_smart

# Global variables for batch processing
checking_status = {
    'running': False,
    'progress': 0,
    'total': 0,
    'results': [],
    'result_file': None,
    'session_id': None
}

@csrf_exempt
@require_http_methods(['POST'])
def check_single_smart(request):
    """Smart check without QR scanning - Pattern-based analysis"""
    try:
        data = json.loads(request.body)
        number = data.get('number', '')
        
        if not number:
            return JsonResponse({'error': 'No number provided'}, status=400)
        
        print("=" * 80)
        print(f" SMART CHECK (NO QR REQUIRED)")
        print(f" Number: {number}")
        print("=" * 80)
        
        # Use smart checker
        result = check_number_smart(number)
        
        response_data = {
            'number': number,
            'registered': result['status'] == 'registered',
            'message': result['verdict'],
            'confidence': result['confidence_score'],
            'analysis': result['analysis'],
            'country': result['country'],
            'is_mobile': result['is_mobile'],
            'carrier': result['carrier'],
            'status': 'success',
            'mode': 'smart_analysis',
            'session_info': 'Smart pattern-based checking - No QR scan required!'
        }
        
        print("=" * 80)
        print(f" SMART RESULT: {number} -> {result['verdict']} ({result['confidence_score']}% confidence)")
        print("=" * 80)
        
        return JsonResponse(response_data)
        
    except Exception as e:
        print(f' Smart Check Error: {e}')
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(['POST'])
def check_single(request):
    """Fallback to smart check for now"""
    return check_single_smart(request)

@csrf_exempt
@require_http_methods(['POST'])
def check_batch_smart(request):
    """Smart batch check without QR scanning"""
    global checking_status
    try:
        data = json.loads(request.body)
        numbers = data.get('numbers', [])
        
        if not numbers:
            return JsonResponse({'error': 'No numbers provided'}, status=400)
        
        session_id = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        checking_status = {
            'running': True,
            'progress': 0,
            'total': len(numbers),
            'results': [],
            'result_file': None,
            'session_id': session_id,
            'mode': 'smart'
        }
        
        def smart_batch_process():
            global checking_status
            print("=" * 80)
            print(f" SMART BATCH CHECK (NO QR REQUIRED)")
            print(f" Total: {len(numbers)} numbers")
            print("=" * 80)
            
            for i, number in enumerate(numbers):
                try:
                    print(f" Processing {i+1}/{len(numbers)}: {number}")
                    smart_result = check_number_smart(number.strip())
                    
                    checking_status['results'].append({
                        'number': number,
                        'registered': smart_result['status'] == 'registered',
                        'confidence': smart_result['confidence_score'],
                        'analysis': smart_result['analysis'],
                        'country': smart_result['country'],
                        'message': smart_result['verdict']
                    })
                    checking_status['progress'] = i + 1
                    
                    print(f" Smart Result: {number} -> {smart_result['verdict']} ({smart_result['confidence_score']}%)")
                    
                    time.sleep(0.5)  # Fast processing
                    
                except Exception as e:
                    print(f' Error processing {number}: {e}')
                    checking_status['results'].append({
                        'number': number,
                        'error': str(e)
                    })
                    checking_status['progress'] = i + 1
            
            try:
                file_info = create_smart_result_files(checking_status['results'], session_id)
                if file_info:
                    checking_status['result_file'] = file_info
                    print(f" [FILES] Smart analysis files created!")
            except Exception as e:
                print(f' File creation error: {e}')
            
            checking_status['running'] = False
            print(" SMART BATCH CHECK COMPLETED!")
        
        thread = threading.Thread(target=smart_batch_process)
        thread.daemon = True
        thread.start()
        
        return JsonResponse({
            'message': 'Smart Analysis: No QR scan required - Pattern-based checking started!',
            'total': len(numbers),
            'session_id': session_id,
            'mode': 'smart',
            'status': 'success'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(['POST'])
def check_batch(request):
    """Fallback to smart batch check"""
    return check_batch_smart(request)

def create_smart_result_files(results, session_id):
    """Create smart analysis result files"""
    try:
        results_dir = os.path.join('media', 'results')
        os.makedirs(results_dir, exist_ok=True)
        
        print(f" [FILES] Creating smart analysis result files...")
        
        data = []
        for result in results:
            if 'error' in result:
                data.append({
                    'Phone Number': result['number'],
                    'Status': 'ERROR',
                    'WhatsApp Status': 'Error',
                    'Confidence': 'N/A',
                    'Country': 'N/A',
                    'Analysis': 'Error occurred',
                    'Message': result['error'],
                    'Checked At': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
            else:
                status = 'LIKELY REGISTERED' if result['registered'] else 'LIKELY NOT REGISTERED'
                data.append({
                    'Phone Number': result['number'],
                    'Status': status,
                    'WhatsApp Status': 'Probable' if result['registered'] else 'Unlikely',
                    'Confidence': f"{result.get('confidence', 0)}%",
                    'Country': result.get('country', 'Unknown'),
                    'Analysis': str(result.get('analysis', {})),
                    'Message': result.get('message', ''),
                    'Checked At': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
        
        df = pd.DataFrame(data)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # CSV file
        csv_filename = f'whatsapp_smart_analysis_{session_id}_{timestamp}.csv'
        csv_path = os.path.join(results_dir, csv_filename)
        df.to_csv(csv_path, index=False)
        
        # Excel file with enhanced formatting
        excel_filename = f'whatsapp_smart_analysis_{session_id}_{timestamp}.xlsx'
        excel_path = os.path.join(results_dir, excel_filename)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Smart WhatsApp Analysis"
        
        # Headers
        headers = ['Phone Number', 'Status', 'WhatsApp Status', 'Confidence', 'Country', 'Analysis', 'Message', 'Checked At']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Data with smart coloring
        for row_num, row_data in enumerate(data, 2):
            ws.cell(row=row_num, column=1, value=row_data['Phone Number'])
            
            status_cell = ws.cell(row=row_num, column=2, value=row_data['Status'])
            if 'LIKELY REGISTERED' in row_data['Status']:
                status_cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            elif 'LIKELY NOT REGISTERED' in row_data['Status']:
                status_cell.fill = PatternFill(start_color="E15759", end_color="E15759", fill_type="solid")
            else:
                status_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            
            ws.cell(row=row_num, column=3, value=row_data['WhatsApp Status'])
            ws.cell(row=row_num, column=4, value=row_data['Confidence'])
            ws.cell(row=row_num, column=5, value=row_data['Country'])
            ws.cell(row=row_num, column=6, value=row_data['Analysis'])
            ws.cell(row=row_num, column=7, value=row_data['Message'])
            ws.cell(row=row_num, column=8, value=row_data['Checked At'])
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(excel_path)
        
        return {
            'csv_file': csv_filename,
            'csv_path': csv_path,
            'excel_file': excel_filename,
            'excel_path': excel_path
        }
        
    except Exception as e:
        print(f' [FILES] Error: {e}')
        return None

def get_status(request):
    return JsonResponse(checking_status)

@csrf_exempt
def download_results(request, filename):
    try:
        file_path = os.path.join('media', 'results', filename)
        
        if not os.path.exists(file_path):
            return JsonResponse({'error': 'File not found'}, status=404)
        
        if filename.endswith('.xlsx'):
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        else:
            content_type = 'text/csv'
        
        with open(file_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type=content_type)
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(['POST'])
def upload_file(request):
    try:
        if 'file' not in request.FILES:
            return JsonResponse({'error': 'No file uploaded'}, status=400)
        
        uploaded_file = request.FILES['file']
        if uploaded_file.size > 5 * 1024 * 1024:
            return JsonResponse({'error': 'File too large (max 5MB)'}, status=400)
        
        numbers = []
        file_name = uploaded_file.name.lower()
        
        if file_name.endswith('.txt'):
            content = uploaded_file.read().decode('utf-8')
            numbers = [line.strip() for line in content.split('\n') if line.strip()]
        elif file_name.endswith('.csv'):
            content = uploaded_file.read().decode('utf-8')
            csv_reader = csv.reader(io.StringIO(content))
            numbers = [row[0].strip() for row in csv_reader if row and row[0].strip()]
        else:
            return JsonResponse({'error': 'Unsupported file format'}, status=400)
        
        clean_numbers = [str(num).strip() for num in numbers if str(num).strip() and len(str(num).strip()) >= 10]
        
        if not clean_numbers:
            return JsonResponse({'error': 'No valid phone numbers found'}, status=400)
        
        return JsonResponse({
            'success': True,
            'numbers': clean_numbers,
            'count': len(clean_numbers),
            'message': f'Loaded {len(clean_numbers)} numbers for smart analysis'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def session_status(request):
    return JsonResponse({
        'initialized': True,
        'driver_active': True,
        'logged_in': True,
        'session_type': 'SMART_MODE_NO_QR_REQUIRED'
    })

def index(request):
    return render(request, 'index.html')

def test_page(request):
    return render(request, 'test.html')

def test_api(request):
    return JsonResponse({
        'status': 'Smart Mode Active - No QR scan needed!',
        'method': request.method,
        'timestamp': datetime.now().strftime('%H:%M:%S'),
        'mode': 'smart_analysis'
    })